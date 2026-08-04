#!/usr/bin/env python3
"""
Regenera los JSON de src/data/ desde las fuentes del pipeline Tableau de Explora.

Fuentes (requiere /Volumes/comun montado):
  - Detalle Biodiesel Argentina.hyper  → empresa-mes 2008→hoy (fuente primaria)
  - Cumplimiento Petrolera.hyper       → solicitado/cupo por petrolera-mes
  - Secretaria de Energía.hyper        → secretarios de energía + presidencias
  - Formulas de precio.hyper           → parámetros fórmulas de precio SE
  - Mercado Argentino Derivados Petroleo Table.hyper → ventas GO G2+G3 (corte real)
  - ARGENTINA BIODIESEL MARKET REV FINAL.xlsx → PROD CAPACITY, HOLDING, CAMARAS
  - Lineas para grafico corte obligatorio.xlsx → % corte obligatorio mensual

Uso:
    python3 scripts/regenerate_data.py [--dry-run]

El script NO estima datos: si falta un dato o una validación falla, aborta
con mensaje explícito. Revisar el diff antes de commitear.
"""

import json
import shutil
import sys
import tempfile
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    from tableauhyperapi import Connection, HyperProcess, Telemetry
except ImportError:
    sys.exit("Falta tableauhyperapi:  pip3 install tableauhyperapi")
try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl:  pip3 install openpyxl")

# ---------------------------------------------------------------- configuración

VOL = Path("/Volumes/comun/01. TABLEAU")
SRC = {
    "detalle":     VOL / "EXP MKTS DATABASES/Revision Actual/Detalle Biodiesel Argentina.hyper",
    "cumpli_petro": VOL / "EXP MKTS DATABASES/Revision Actual/Cumplimiento Petrolera.hyper",
    "secretarios": VOL / "EXP MKTS DATABASES/Revision Actual/Secretaria de Energía.hyper",
    "formulas":    VOL / "EXP MKTS DATABASES/Revision Actual/Formulas de precio.hyper",
    "derivados":   VOL / "EXP MKTS DATABASES/Revision Actual/Mercado Argentino Derivados Petroleo Table.hyper",
    "maestro":     VOL / "EXP MKTSCAN - MARKET ANALYSIS/BIODIESEL/Revision Actual/ARGENTINA BIODIESEL MARKET REV FINAL.xlsx",
    "corte_oblig": VOL / "EXP MKTSCAN - DATASOURCES/Revision Actual/Lineas para grafico corte obligatorio.xlsx",
}

OUT_DIR = Path(__file__).resolve().parent.parent / "src" / "data"
DENSIDAD_BIO = 0.885  # ton/m3, factor usado en los workbooks Tableau
# Nombres de exhibición: la fuente SE conserva razones sociales históricas
ALIAS_PETROLERAS = {
    "SHELL ARGENTINA C.A.P.S.A.": "RAIZEN (ex SHELL)",
}

# La SE renombró retroactivamente a ESSO S.A.P.A. como AXION en el dataset de
# biodiesel (todas las compras históricas figuran como AXION), mientras que en
# el de derivados partió los nombres: ESSO hasta 2018-03, AXION desde 2018-04.
# Decisión HDO (2026-08-03): usar ese corte de los datos SE como límite de la
# entidad en TODOS los datasets - hasta 2018-03 inclusive es ESSO S.A.P.A.,
# desde 2018-04 es AXION ENERGY ARGENTINA S.A.
CORTE_ESSO_AXION = "2018-03"  # último mes imputado a ESSO S.A.P.A.
NOMBRE_AXION = "AXION ENERGY ARGENTINA S.A."
NOMBRE_ESSO = "ESSO S.A.P.A."

# El hyper de detalle trae a estas aceiteras como NO INTEGRADA en toda su
# historia; decisión HDO (2026-08-03, coincide con su workbook Tableau):
# son INTEGRADAS.
CATEGORIA_OVERRIDE = {
    "VICENTÍN S.A.I.C.": "INTEGRADA",
    "NOBLE ARGENTINA S.A.": "INTEGRADA",
}

# Nombres de exhibición de elaboradoras (lista HDO 2026-08-04). Se aplican en
# TODAS las fuentes que traen el nombre (detalle, capacidad, holding, cámaras)
# para que los joins internos sigan cerrando.
ALIAS_EMPRESAS = {
    "MOLINOS RÍO DE LA PLATA S.A.": "MOLINOS S.A.",
    "ACEITE GENERAL DEHEZA S.A.": "AGD S.A.",
    "NOBLE ARGENTINA S.A.": "NOBLE S.A.",
    "PATAGONIA BIOENERGIA S.A.": "PATAGONIA S.A.",
    "ADVANCED ORGANIC MATERIALS S.A.": "AOM S.A.",
    "ENERGÍA RENOVABLE S.A.": "ENRESA",
    "ENERGÍAS RENOVABLES ARGENTINA S.A.": "ERA S.A.",
    "DOBLE L BIOENERGIAS S.A.": "DOBLE L S.A.",
}


def alias_empresa(nombre):
    return ALIAS_EMPRESAS.get(nombre, nombre)


# Correcciones de grupo económico sobre el hyper (decisiones HDO 2026-08-03/04):
# - Biomadero no pertenece al GRUPO BOJANICH, queda independiente.
# - Bio Nogoya y Héctor A. Bolzan y Cía. constituyen el GRUPO BOLZÁN.
GRUPO_OVERRIDE = {
    "BIOMADERO S.A.": "BIOMADERO S.A.",
    "BIO NOGOYA S.A.": "GRUPO BOLZÁN",
    "HÉCTOR A. BOLZAN Y CÍA. S.R.L.": "GRUPO BOLZÁN",
}

# Renombres de grupos económicos (decisión HDO 2026-08-04)
ALIAS_GRUPOS = {
    "GRUPO ROSARIO BIO": "GRUPO PUCCIARIELLO",
}

# Holdings de holdings (decisión HDO 2026-08-04): el GRUPO ESSENTIAL ENERGY
# está conformado por dos grupos - GRUPO PUCCIARIELLO (ex Rosario Bio) y
# GRUPO BOLZÁN.
SUPERGRUPOS = {
    "GRUPO PUCCIARIELLO": "GRUPO ESSENTIAL ENERGY",
    "GRUPO BOLZÁN": "GRUPO ESSENTIAL ENERGY",
}


def alias_petrolera(nombre, fecha=None):
    nombre = ALIAS_PETROLERAS.get(nombre, nombre)
    if fecha is not None:
        if nombre == NOMBRE_AXION and fecha <= CORTE_ESSO_AXION:
            return NOMBRE_ESSO
        if nombre == NOMBRE_ESSO and fecha > CORTE_ESSO_AXION:
            return NOMBRE_AXION
    return nombre


PETROLERAS_COLS = [
    "AXION ENERGY ARGENTINA S.A.", "DESTILERÍA ARGENTINA DE PETRÓLEO S.A.",
    "PETROBRAS ARGENTINA S.A.", "PETROLERA DEL CONOSUR S.A.", "POLIPETROL S.A.",
    "REFINERÍA DEL NORTE S.A.", "SHELL ARGENTINA C.A.P.S.A.", "YPF S.A.",
    "OIL COMBUSTIBLES S.A.", "NEW AMERICAN OIL S.A.", "REFI PAMPA S.A.",
    "TRAFIGURA ARGENTINA S.A.", "ENERGÍA ARGENTINA S.A. (ENARSA)",
    "ENERGÍA Y DERIVADOS DEL PETRÓLEO S.A.", "PETROIL S.A.",
    "DIVERSE FUELS S.A.", "PETROLERA DEGAB S.A.",
]

# ------------------------------------------------------------------- utilidades

def ym(d):
    return f"{d.year:04d}-{d.month:02d}"

def r1(x):
    return round(x, 1) if x is not None else None

def r4(x):
    return round(x, 4) if x is not None else None

def nfc(s):
    return unicodedata.normalize("NFC", s) if isinstance(s, str) else s

def fail(msg):
    sys.exit(f"\n✗ ERROR: {msg}\nNo se escribió ningún archivo.")

def check(cond, msg):
    if not cond:
        fail(msg)

WARNINGS = []
def warn(msg):
    WARNINGS.append(msg)
    print(f"  ⚠ {msg}")


class Hyper:
    """Copia los .hyper a un directorio temporal (lectura segura desde red)."""

    def __init__(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="regen_hyper_"))
        self.hp = HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU)

    def query(self, name, sql):
        local = self.tmp / f"{name}.hyper"
        if not local.exists():
            shutil.copy(SRC[name], local)
        with Connection(endpoint=self.hp.endpoint, database=str(local)) as conn:
            return conn.execute_list_query(sql)

    def close(self):
        self.hp.close()
        shutil.rmtree(self.tmp, ignore_errors=True)


def leer_excel(path, hoja_buscada):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hoja = next((s for s in wb.sheetnames if nfc(s).strip() == hoja_buscada), None)
    check(hoja is not None, f"No existe la hoja '{hoja_buscada}' en {path.name}")
    rows = [r for r in wb[hoja].iter_rows(values_only=True)]
    wb.close()
    return rows

# ------------------------------------------------------------------ extracción

def extraer_detalle(hy):
    cols = (
        '"DATE", "EMPRESA ELABORADORA", "CATEGORIA", "GRUPO ECONÓMICO", '
        '"PROVINCIA", "LOCALIDAD", "Is our Company Flag", '
        '"PRODUCTION [ton]", "BIODIESEL QUOTA [ton]", "BIODIESEL QUOTA SALES [ton]", '
        '"BIODIESEL XQUOTA SALES [ton]", "BIODIESEL COTAB SALES [ton]", '
        '"BIODIESEL EXPORTS [ton]", '
        + ", ".join(f'"{c}"' for c in PETROLERAS_COLS)
    )
    rows = hy.query("detalle", f'SELECT {cols} FROM "Extract"."Extract" ORDER BY "DATE"')
    registros = []
    for r in rows:
        fecha = ym(r[0])
        empresa = nfc(r[1])  # nombre crudo: clave de los overrides
        grupo = alias_empresa(GRUPO_OVERRIDE.get(empresa, nfc(r[3])))
        grupo = ALIAS_GRUPOS.get(grupo, grupo)
        d = dict(
            fecha=fecha, empresa=alias_empresa(empresa),
            categoria=CATEGORIA_OVERRIDE.get(empresa, r[2]),
            grupo=grupo, supergrupo=SUPERGRUPOS.get(grupo),
            provincia=r[4], localidad=r[5], explora=(r[6] == "Y"),
            prod=r[7] or 0, cupo=r[8] or 0, vc=r[9] or 0,
            xq=r[10] or 0, cotab=r[11] or 0, exp=r[12] or 0,
            petroleras={alias_petrolera(PETROLERAS_COLS[i], fecha): v
                        for i, v in enumerate(r[13:]) if v},
        )
        check(d["categoria"] in ("INTEGRADA", "NO INTEGRADA", "COMERCIALIZADORA"),
              f"Empresa sin categoría válida: {d['empresa']} {d['fecha']} → {d['categoria']!r}")
        registros.append(d)
    return registros


# Destinos de gas oil que no llevan corte obligatorio: quedan fuera del
# denominador del % de corte real (calibrado contra la serie de explorarg:
# con estas exclusiones 2010-2015 y 2024-2025 coinciden a ±0,05pp).
# Por decisión de HDO (2026-08-03), Estado y S/N SÍ se incluyen en el denominador.
SECTORES_SIN_CORTE = ("Bunker Cabotaje", "Bunker Internacional", "Usinas Eléctricas")


# Nombre de cada petrolera en el dataset de derivados (SE) → nombre en el
# detalle de biodiesel. Solo mezcladoras relevantes con cupo asignado.
MAPA_PETROLERAS_GO = {
    "YPF S.A.": "YPF S.A.",
    "SHELL C.A.P.S.A.": "RAIZEN (ex SHELL)",
    "AXION S.A.": "AXION ENERGY ARGENTINA S.A.",
    "ESSO S.A.P.A.": "ESSO S.A.P.A.",
    "TRAFIGURA S.A.": "TRAFIGURA ARGENTINA S.A.",
    "DAPSA S.A.": "DESTILERÍA ARGENTINA DE PETRÓLEO S.A.",
    "REFIPAMPA S.A.": "REFI PAMPA S.A.",
    "REFINOR S.A.": "REFINERÍA DEL NORTE S.A.",
    "NEW AMERICAN OIL": "NEW AMERICAN OIL S.A.",
    "PETROLERA DEGAB S.A.": "PETROLERA DEGAB S.A.",
}


def extraer_go_todas_empresas(hy):
    """GO G2+G3 (m3) por empresa-mes para TODAS las vendedoras del dataset,
    sin sectores exentos. Alimenta el cuadro de demanda de bio por petrolera:
    las que venden GO sin comprar biodiesel también deben aparecer."""
    rows = hy.query("derivados", f'''
        SELECT "FECHA", "empresa",
               SUM(COALESCE("Gasoil Grado 2 (Común)", 0) + COALESCE("Gasoil Grado 3 (Ultra)", 0))
        FROM "Extract"."Extract"
        WHERE "FECHA" IS NOT NULL AND "empresa" IS NOT NULL
          AND "SECTOR" NOT IN {SECTORES_SIN_CORTE!r}
        GROUP BY 1, 2 ORDER BY 1''')
    out = defaultdict(dict)
    for f, emp, v in rows:
        if v and v > 0:
            fecha = ym(f)
            nombre = alias_petrolera(MAPA_PETROLERAS_GO.get(emp, emp.strip()), fecha)
            out[fecha][nombre] = round((out[fecha].get(nombre, 0) or 0) + v, 1)
    return [dict(fecha=f, **vals) for f, vals in sorted(out.items())]


def extraer_go_mensual(hy):
    """Ventas GO Grado 2 + Grado 3 (m3) por mes, sin sectores exentos de corte."""
    rows = hy.query("derivados", f'''
        SELECT "FECHA",
               SUM(COALESCE("Gasoil Grado 2 (Común)", 0)),
               SUM(COALESCE("Gasoil Grado 3 (Ultra)", 0))
        FROM "Extract"."Extract"
        WHERE "SECTOR" NOT IN {SECTORES_SIN_CORTE!r}
        GROUP BY 1 ORDER BY 1''')
    return {ym(r[0]): {"go2": r[1], "go3": r[2]} for r in rows if r[0]}


def normalizar_sector(s):
    """El dataset SE trae sectores con capitalización inconsistente según la
    época ('transporte Público de Pasajeros' vs 'Transporte...'): se unifican."""
    s = (s or "S/N").strip()
    return s[0].upper() + s[1:]


def extraer_go_sectores(hy):
    """GO G2+G3 (m3) por sector-mes: total del sistema y por petrolera mapeada.
    Sin exclusiones — el filtro de sectores se aplica en el front."""
    tot_rows = hy.query("derivados", '''
        SELECT "FECHA", "SECTOR",
               SUM(COALESCE("Gasoil Grado 2 (Común)", 0) + COALESCE("Gasoil Grado 3 (Ultra)", 0))
        FROM "Extract"."Extract" WHERE "FECHA" IS NOT NULL
        GROUP BY 1, 2 ORDER BY 1''')
    total = defaultdict(dict)
    sectores = set()
    for f, sector, v in tot_rows:
        s = normalizar_sector(sector)
        sectores.add(s)
        total[ym(f)][s] = round((total[ym(f)].get(s, 0) or 0) + v, 1)

    pet_rows = hy.query("derivados", f'''
        SELECT "FECHA", "empresa", "SECTOR",
               SUM(COALESCE("Gasoil Grado 2 (Común)", 0) + COALESCE("Gasoil Grado 3 (Ultra)", 0))
        FROM "Extract"."Extract"
        WHERE "FECHA" IS NOT NULL AND "empresa" IN {tuple(MAPA_PETROLERAS_GO)!r}
        GROUP BY 1, 2, 3 ORDER BY 1''')
    por_pet = defaultdict(lambda: defaultdict(dict))
    for f, emp, sector, v in pet_rows:
        s = normalizar_sector(sector)
        pet = alias_petrolera(MAPA_PETROLERAS_GO[emp], ym(f))
        prev = por_pet[ym(f)][pet].get(s, 0) or 0
        por_pet[ym(f)][pet][s] = round(prev + v, 1)

    return dict(
        sectores=sorted(sectores),
        mensual_total=[dict(fecha=f, **vals) for f, vals in sorted(total.items())],
        mensual_petrolera=[
            dict(fecha=f, petrolera=p, **vals)
            for f, pets in sorted(por_pet.items()) for p, vals in sorted(pets.items())
        ],
    )


def extraer_go_por_petrolera(hy):
    """GO G2+G3 (m3) por petrolera-mes, con el nombre del dataset de biodiesel."""
    rows = hy.query("derivados", f'''
        SELECT "FECHA", "empresa",
               SUM(COALESCE("Gasoil Grado 2 (Común)", 0) + COALESCE("Gasoil Grado 3 (Ultra)", 0))
        FROM "Extract"."Extract"
        WHERE "SECTOR" NOT IN {SECTORES_SIN_CORTE!r}
          AND "empresa" IN {tuple(MAPA_PETROLERAS_GO)!r}
        GROUP BY 1, 2 ORDER BY 1''')
    out = defaultdict(dict)
    for r in rows:
        if r[0]:
            fecha = ym(r[0])
            pet = alias_petrolera(MAPA_PETROLERAS_GO[r[1]], fecha)
            out[fecha][pet] = round((out[fecha].get(pet, 0) or 0) + r[2], 1)
    return dict(out)


def extraer_corte_obligatorio():
    rows = leer_excel(SRC["corte_oblig"], "%Corte Obligatorio")
    out = {}
    for r in rows[1:]:
        if r and r[0] and isinstance(r[0], datetime):
            out[ym(r[0])] = float(r[1])
    check(len(out) > 150, f"Serie de corte obligatorio sospechosamente corta: {len(out)} meses")
    return out


def extraer_secretarios(hy):
    rows = hy.query("secretarios", '''
        SELECT DISTINCT "NUMERO", "NOMBRE SEC", "CARGO", "PRESIDENTE",
                        "INGRESO", "EGRESO", "PARTIDO POLITICO", "COALICION"
        FROM "Extract"."Extract"
        WHERE "NOMBRE SEC" IS NOT NULL ORDER BY "INGRESO"''')
    return [dict(n=r[0], nombre=r[1], cargo=r[2], presidente=r[3],
                 desde=str(r[4]), hasta=str(r[5]) if r[5] else None,
                 partido=r[6], coalicion=r[7]) for r in rows]


def extraer_subsecretarios(hy):
    """Subsecretarios/funcionarios del área hidrocarburos con injerencia en biodiesel."""
    rows = hy.query("secretarios", '''
        SELECT DISTINCT "Sub Secretário de Hidrocarburos", "Cargo", "Nombramiento",
                        "Inicio", "Final", "Presidente"
        FROM "Extract"."Extract"
        WHERE "Sub Secretário de Hidrocarburos" IS NOT NULL
        ORDER BY "Inicio"''')
    return [dict(nombre=r[0], cargo=r[1], nombramiento=r[2],
                 desde=str(r[3]), hasta=str(r[4]) if r[4] else None,
                 presidente=r[5]) for r in rows]


def extraer_formulas(hy):
    rows = hy.query("formulas", '''
        SELECT "Date", "Fórmula de precio", "Fecha Inicio Fórmula",
               "% Corte Obligatorio", "Consumo de aceite", "Costo Adquisición Aceite",
               "Retorno %", "Indice IPC Cobertura Nacional", "Infacion Argentina"
        FROM "Extract"."Extract" WHERE "Date" IS NOT NULL ORDER BY "Date"''')
    serie = [dict(fecha=ym(r[0]), formula=r[1], inicio=str(r[2]) if r[2] else None,
                  corte=r[3], consumo_aceite=r[4], costo_aceite=r[5], retorno=r[6])
             for r in rows]
    periodos = []
    for s in serie:
        if not periodos or periodos[-1]["formula"] != s["formula"]:
            periodos.append(dict(formula=s["formula"], desde=s["fecha"]))
    return serie, periodos


def extraer_cumpli_petro(hy):
    rows = hy.query("cumpli_petro", '''
        SELECT "MES", "OIL COMPANIES-1", SUM(COALESCE("VOLUMEN SOLICITADA",0)),
               SUM(COALESCE("CUOTA",0))
        FROM "Extract"."Extract" WHERE "MES" IS NOT NULL
        GROUP BY 1,2 ORDER BY 1''')
    return [dict(fecha=ym(r[0]), petrolera=alias_petrolera(nfc(r[1]), ym(r[0])),
                 solicitado=r1(r[2]), cupo=r1(r[3]))
            for r in rows]


def extraer_capacidad():
    rows = leer_excel(SRC["maestro"], "PROD CAPACITY")
    serie = [dict(fecha=ym(r[0]), empresa=alias_empresa(nfc(str(r[1]).strip())),
                  capacidad=float(r[2]), condicion=r[3])
             for r in rows[1:] if r and r[0] and r[1] and r[2] is not None]
    hrows = leer_excel(SRC["maestro"], "HOLDING")
    plantas = []
    for r in hrows[1:]:
        if not (r and r[0]):
            continue
        lat, lng = r[6], r[7]
        plantas.append(dict(
            empresa=alias_empresa(nfc(str(r[0]).strip())),
            holding=alias_empresa(nfc(str(r[1]).strip())) if r[1] else None,
            capacidad=float(r[2]) if r[2] is not None else None,
            segmento=r[3], grupo=r[4],
            lat=float(lat) if lat not in (None, "") else None,
            lng=float(lng) if lng not in (None, "") else None,
        ))
    return serie, plantas


def extraer_camaras():
    rows = leer_excel(SRC["maestro"], "CAMARAS")
    out = {}
    for r in rows[1:]:
        if r and r[0] and r[1]:
            out[(ym(r[0]), alias_empresa(nfc(str(r[1]).strip())))] = dict(
                camara=r[2], camara_actual=r[5] if len(r) > 5 else None)
    return out

# ---------------------------------------------------------------- agregaciones

def agregar(registros, go_mensual, corte_oblig):
    meses = sorted({d["fecha"] for d in registros})
    ultimo = meses[-1]

    mensual = defaultdict(lambda: defaultdict(float))
    m_cat = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for d in registros:
        m = mensual[d["fecha"]]
        for k in ("prod", "cupo", "vc", "xq", "cotab", "exp"):
            m[k] += d[k]
        c = m_cat[d["fecha"]][d["categoria"]]
        for k in ("prod", "cupo", "vc", "exp"):
            c[k] += d[k]

    serie_mensual = [dict(fecha=f, **{k: r1(v) for k, v in mensual[f].items()}) for f in meses]

    anual = defaultdict(lambda: defaultdict(float))
    a_cat = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for d in registros:
        a = anual[d["fecha"][:4]]
        for k in ("prod", "cupo", "vc", "xq", "cotab", "exp"):
            a[k] += d[k]
        c = a_cat[d["fecha"][:4]][d["categoria"]]
        for k in ("prod", "cupo", "vc", "exp"):
            c[k] += d[k]
    serie_anual = [dict(anio=int(y), **{k: r1(v) for k, v in anual[y].items()})
                   for y in sorted(anual)]
    serie_anual_cat = [dict(anio=int(y), categoria=c, **{k: r1(v) for k, v in vals.items()})
                       for y in sorted(a_cat) for c, vals in sorted(a_cat[y].items())]
    serie_mensual_cat = [dict(fecha=f, categoria=c, **{k: r1(v) for k, v in vals.items()})
                         for f in meses for c, vals in sorted(m_cat[f].items())]

    # corte real: bio m3 / (GO2+GO3 m3)
    corte_mensual = []
    for f in meses:
        go = go_mensual.get(f)
        oblig = corte_oblig.get(f)
        if go is None or (go["go2"] + go["go3"]) <= 0:
            if f >= "2010-01":
                warn(f"Sin ventas GO para {f}: corte real de ese mes queda nulo")
            continue
        go_m3 = go["go2"] + go["go3"]
        bio_m3 = mensual[f]["vc"] / DENSIDAD_BIO
        corte_mensual.append(dict(
            fecha=f, obligatorio=r4(oblig), real=r4(bio_m3 / go_m3),
            bio_m3=r1(bio_m3), go_m3=r1(go_m3),
            deficit_ton=r1(max(0.0, (oblig - bio_m3 / go_m3)) * go_m3 * DENSIDAD_BIO)
            if oblig is not None else None,
        ))

    corte_anual = []
    por_anio = defaultdict(list)
    for c in corte_mensual:
        por_anio[c["fecha"][:4]].append(c)
    for y in sorted(por_anio):
        cs = [c for c in por_anio[y] if c["obligatorio"] is not None]
        if not cs:
            continue
        go_tot = sum(c["go_m3"] for c in cs)
        bio_tot = sum(c["bio_m3"] for c in cs)
        corte_anual.append(dict(
            anio=int(y),
            obligatorio=r4(sum(c["obligatorio"] * c["go_m3"] for c in cs) / go_tot),
            real=r4(bio_tot / go_tot),
            deficit_ton=r1(sum(c["deficit_ton"] or 0 for c in cs)),
            meses=len(cs),
        ))

    return dict(ultimo=ultimo, meses=meses, mensual=serie_mensual, anual=serie_anual,
                anual_categoria=serie_anual_cat, mensual_categoria=serie_mensual_cat,
                corte_mensual=corte_mensual, corte_anual=corte_anual)


def agregar_empresas(registros, camaras):
    por_emp = defaultdict(list)
    for d in registros:
        por_emp[d["empresa"]].append(d)
    empresas = []
    for nombre, rows in sorted(por_emp.items()):
        rows.sort(key=lambda d: d["fecha"])
        last = rows[-1]
        cam = camaras.get((last["fecha"], nombre)) or {}
        empresas.append(dict(
            empresa=nombre, categoria=last["categoria"], grupo=last["grupo"],
            supergrupo=last["supergrupo"],
            provincia=last["provincia"], localidad=last["localidad"],
            explora=last["explora"], camara=cam.get("camara_actual") or cam.get("camara"),
            serie=[[d["fecha"], r1(d["prod"]), r1(d["cupo"]), r1(d["vc"]),
                    r1(d["xq"]), r1(d["exp"]), r1(d["cotab"])] for d in rows],
        ))
    return empresas


def agregar_petroleras(registros, cumpli):
    mensual = defaultdict(lambda: defaultdict(float))
    matriz_12m = defaultdict(lambda: defaultdict(float))
    meses = sorted({d["fecha"] for d in registros})
    ult12 = set(meses[-12:])
    for d in registros:
        for p, v in d["petroleras"].items():
            mensual[d["fecha"]][p] += v
            if d["fecha"] in ult12:
                matriz_12m[d["empresa"]][p] += v
    serie = [dict(fecha=f, **{p: r1(v) for p, v in sorted(mensual[f].items())})
             for f in sorted(mensual)]
    matriz = [dict(empresa=e, ventas={p: r1(v) for p, v in sorted(vs.items())})
              for e, vs in sorted(matriz_12m.items())]
    return dict(mensual=serie, matriz_12m=matriz, cumplimiento=cumpli,
                periodo_12m={"desde": min(ult12), "hasta": max(ult12)})

def generar_dashboard_legacy(registros):
    """dashboard.json con el esquema original de la Portada, con datos frescos.
    Mismos nombres de campo que la v2 generada en claude.ai: la Portada no
    necesita cambios de código para actualizarse."""
    meses = sorted({d["fecha"] for d in registros})
    ult12 = set(meses[-12:])

    mensual = defaultdict(lambda: defaultdict(float))
    anual = defaultdict(lambda: defaultdict(float))
    cat_anual = defaultdict(lambda: defaultdict(float))
    emp12 = defaultdict(lambda: defaultdict(float))
    emp_cat = {}
    pet12 = defaultdict(float)
    prov12 = defaultdict(float)
    grp12 = defaultdict(lambda: defaultdict(float))
    grp_cat = {}

    for d in registros:
        m = mensual[d["fecha"]]
        m["PRODUCTION [ton]"] += d["prod"]
        m["BIODIESEL QUOTA [ton]"] += d["cupo"]
        m["BIODIESEL QUOTA SALES [ton]"] += d["vc"]
        m["BIODIESEL XQUOTA SALES [ton]"] += d["xq"]
        m["BIODIESEL EXPORTS [ton]"] += d["exp"]

        a = anual[d["fecha"][:4]]
        a["PRODUCTION [ton]"] += d["prod"]
        a["BIODIESEL QUOTA [ton]"] += d["cupo"]
        a["BIODIESEL QUOTA SALES [ton]"] += d["vc"]
        a["BIODIESEL EXPORTS [ton]"] += d["exp"]

        cat_anual[d["fecha"][:4]][d["categoria"]] += d["prod"]

        if d["fecha"] in ult12:
            e = emp12[d["empresa"]]
            e["PRODUCTION [ton]"] += d["prod"]
            e["BIODIESEL QUOTA SALES [ton]"] += d["vc"]
            e["BIODIESEL EXPORTS [ton]"] += d["exp"]
            e["BIODIESEL QUOTA [ton]"] += d["cupo"]
            emp_cat[d["empresa"]] = d["categoria"]
            for pet, v in d["petroleras"].items():
                pet12[pet] += v
            if d["provincia"]:
                prov12[d["provincia"]] += d["prod"]
            g = grp12[d["grupo"] or d["empresa"]]
            g["PRODUCTION [ton]"] += d["prod"]
            g["BIODIESEL QUOTA SALES [ton]"] += d["vc"]
            grp_cat[d["grupo"] or d["empresa"]] = d["categoria"]

    top = sorted(emp12.items(), key=lambda kv: -kv[1]["PRODUCTION [ton]"])[:20]
    return dict(
        meta=dict(
            ultimo_mes=meses[-1], primer_mes=meses[0],
            total_empresas=len({d["empresa"] for d in registros}),
            total_petroleras_mezcladoras=len(PETROLERAS_COLS),
            correccion_aplicada="Patagonia Bioenergia S.A. reclasificada como INTEGRADA",
        ),
        mensual=[dict(FECHA=f, **{k: r1(v) for k, v in vals.items()})
                 for f, vals in sorted(mensual.items())],
        anual=[dict(AÑO=int(y), **{k: r1(v) for k, v in vals.items()})
               for y, vals in sorted(anual.items())],
        categoria_anual=[
            dict(AÑO=int(y), COMERCIALIZADORA=r1(vals.get("COMERCIALIZADORA", 0)),
                 INTEGRADA=r1(vals.get("INTEGRADA", 0)),
                 **{"NO INTEGRADA": r1(vals.get("NO INTEGRADA", 0))})
            for y, vals in sorted(cat_anual.items())
        ],
        top_empresas_12m=[
            dict(
                **{"EMPRESA ELABORADORA": emp, "CATEGORIA": emp_cat[emp]},
                **{k: r1(v) for k, v in vals.items()},
                **{"CUMPLIMIENTO %": r1(vals["BIODIESEL QUOTA SALES [ton]"]
                                        / vals["BIODIESEL QUOTA [ton]"] * 100)
                   if vals["BIODIESEL QUOTA [ton]"] > 0 else None},
            )
            for emp, vals in top
        ],
        ventas_petroleras_12m=[
            dict(PETROLERA=p, TONELADAS=r1(v))
            for p, v in sorted(pet12.items(), key=lambda kv: -kv[1]) if v > 0
        ],
        provincia_12m=[
            dict(PROVINCIA=p, **{"PRODUCTION [ton]": r1(v)})
            for p, v in sorted(prov12.items(), key=lambda kv: -kv[1]) if v > 0
        ],
        grupos_12m=[
            dict(**{"GRUPO ECONÓMICO": g, "CATEGORIA": grp_cat[g]},
                 **{k: r1(v) for k, v in vals.items()})
            for g, vals in sorted(grp12.items(),
                                  key=lambda kv: -kv[1]["PRODUCTION [ton]"])
            if vals["PRODUCTION [ton]"] > 0 or vals["BIODIESEL QUOTA SALES [ton]"] > 0
        ],
        periodo_12m=dict(desde=min(ult12), hasta=max(ult12)),
    )


# ---------------------------------------------------------------- validaciones

def validar(agg, empresas):
    a2025 = next((a for a in agg["corte_anual"] if a["anio"] == 2025), None)
    check(a2025 is not None, "No hay corte anual 2025 para validar")
    check(abs(a2025["real"] - 0.058) < 0.002,
          f"Corte real 2025 = {a2025['real']:.4f}, esperado ≈ 0.058 (cifra ancla)")
    check(abs(a2025["obligatorio"] - 0.075) < 0.003,
          f"Corte obligatorio 2025 = {a2025['obligatorio']:.4f}, esperado ≈ 0.075")

    explora = next((e for e in empresas if e["explora"]), None)
    check(explora is not None, "No se encontró Explora (Is our Company Flag = Y)")
    vc25 = sum(r[3] for r in explora["serie"] if r[0].startswith("2025"))
    cupo25 = sum(r[2] for r in explora["serie"] if r[0].startswith("2025"))
    check(abs(vc25 - 51071) / 51071 < 0.02,
          f"Explora ventas corte 2025 = {vc25:.0f} ton, esperado ≈ 51.071")
    check(cupo25 > 0 and abs(vc25 / cupo25 - 1.22) < 0.05,
          f"Explora cumplimiento 2025 = {vc25 / cupo25:.2%}, esperado ≈ 122%")

    # continuidad de meses
    meses = agg["meses"]
    for a, b in zip(meses, meses[1:]):
        ya, ma = int(a[:4]), int(a[5:])
        esperado = f"{ya + (ma // 12):04d}-{(ma % 12) + 1:02d}"
        check(b == esperado, f"Hueco en la serie mensual: {a} → {b}")

    print(f"  ✓ Corte 2025: real {a2025['real']:.1%} / obligatorio {a2025['obligatorio']:.1%}")
    print(f"  ✓ ESSO/AXION particionadas en {CORTE_ESSO_AXION}")
    print(f"  ✓ Explora 2025: {vc25:,.0f} ton, cumplimiento {vc25 / cupo25:.0%}")
    print(f"  ✓ Serie mensual continua: {meses[0]} → {meses[-1]} ({len(meses)} meses)")

# ---------------------------------------------------------------------- salida

def escribir(nombre, payload, fuentes, dry):
    payload["meta"] = dict(
        **payload.get("meta", {}),
        generado=datetime.now().strftime("%Y-%m-%d %H:%M"),
        fuentes=fuentes,
        nota="Generado por scripts/regenerate_data.py - no editar a mano",
    )
    destino = OUT_DIR / nombre
    if dry:
        print(f"  [dry-run] {nombre}: {len(json.dumps(payload))//1024} KB")
        return
    destino.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
                       encoding="utf-8")
    print(f"  ✓ {nombre} ({destino.stat().st_size // 1024} KB)")


def main():
    dry = "--dry-run" in sys.argv
    check(VOL.exists(), "El volumen /Volumes/comun no está montado")
    for k, p in SRC.items():
        check(p.exists(), f"Fuente '{k}' no encontrada: {p}")

    print("Leyendo fuentes…")
    hy = Hyper()
    try:
        registros = extraer_detalle(hy)
        go_mensual = extraer_go_mensual(hy)
        go_petrolera = extraer_go_por_petrolera(hy)
        go_todas = extraer_go_todas_empresas(hy)
        go_sectores = extraer_go_sectores(hy)
        corte_oblig = extraer_corte_obligatorio()
        secretarios = extraer_secretarios(hy)
        subsecretarios = extraer_subsecretarios(hy)
        form_serie, form_periodos = extraer_formulas(hy)
        cumpli = extraer_cumpli_petro(hy)
        cap_serie, plantas = extraer_capacidad()
        camaras = extraer_camaras()
    finally:
        hy.close()
    print(f"  detalle: {len(registros)} filas empresa-mes · GO: {len(go_mensual)} meses · "
          f"secretarios: {len(secretarios)} · capacidad: {len(cap_serie)} filas")

    print("Agregando…")
    agg = agregar(registros, go_mensual, corte_oblig)
    empresas = agregar_empresas(registros, camaras)
    petroleras = agregar_petroleras(registros, cumpli)

    print("Validando contra cifras ancla…")
    validar(agg, empresas)

    # Consistencia de la partición ESSO/AXION en todas las series con nombre+mes
    def _check_particion(series, origen):
        for row in series:
            f = row["fecha"]
            nombres = [row["petrolera"]] if "petrolera" in row else \
                      [k for k in row if k != "fecha"]
            for n in nombres:
                check(not (n == NOMBRE_AXION and f <= CORTE_ESSO_AXION),
                      f"{origen}: AXION en {f}, debería ser ESSO (corte {CORTE_ESSO_AXION})")
                check(not (n == NOMBRE_ESSO and f > CORTE_ESSO_AXION),
                      f"{origen}: ESSO en {f}, debería ser AXION (corte {CORTE_ESSO_AXION})")
    _check_particion(petroleras["mensual"], "bio compras mensual")
    _check_particion(cumpli, "cumplimiento petrolera")
    _check_particion(go_todas, "GO todas las vendedoras")
    _check_particion([dict(fecha=f, **v) for f, v in go_petrolera.items()],
                     "GO por petrolera")
    _check_particion(go_sectores["mensual_petrolera"], "GO por sector/petrolera")

    # Fechas de asunción presidenciales (registro público) para las bandas de
    # los charts; partido/coalición se toma de los secretarios de cada gestión.
    RANGOS_PRESIDENCIAS = [
        ("Néstor Kirchner", "2003-05-25", "2007-12-10"),
        ("Cristina Fernández de Kirchner", "2007-12-10", "2015-12-10"),
        ("Mauricio Macri", "2015-12-10", "2019-12-10"),
        ("Alberto Fernández", "2019-12-10", "2023-12-10"),
        ("Javier Milei", "2023-12-10", None),
    ]
    coalicion_por_pres = {}
    for s in secretarios:
        if s["presidente"] and s["coalicion"]:
            coalicion_por_pres.setdefault(s["presidente"], s["coalicion"])
    presidencias = [
        dict(presidente=p, desde=d, hasta=h, coalicion=coalicion_por_pres.get(p))
        for p, d, h in RANGOS_PRESIDENCIAS
    ]

    print("Escribiendo JSON…" if not dry else "Simulando escritura…")
    escribir("mercado.json", dict(
        ultimo_mes=agg["ultimo"], mensual=agg["mensual"], anual=agg["anual"],
        anual_categoria=agg["anual_categoria"], mensual_categoria=agg["mensual_categoria"],
    ), ["Detalle Biodiesel Argentina.hyper"], dry)
    escribir("empresas.json", dict(ultimo_mes=agg["ultimo"], empresas=empresas),
             ["Detalle Biodiesel Argentina.hyper", "CAMARAS (Excel maestro)"], dry)
    petroleras["go_mensual"] = [
        dict(fecha=f, **vals) for f, vals in sorted(go_petrolera.items())
    ]
    petroleras["go_empresas_mensual"] = go_todas
    escribir("petroleras.json", petroleras,
             ["Detalle Biodiesel Argentina.hyper", "Cumplimiento Petrolera.hyper",
              "Mercado Argentino Derivados Petroleo Table.hyper"], dry)
    escribir("capacidad.json", dict(serie=cap_serie, plantas=plantas),
             ["PROD CAPACITY + HOLDING (Excel maestro)"], dry)
    escribir("corte.json", dict(
        mensual=agg["corte_mensual"], anual=agg["corte_anual"],
        densidad_bio=DENSIDAD_BIO,
    ), ["Lineas para grafico corte obligatorio.xlsx",
        "Mercado Argentino Derivados Petroleo Table.hyper",
        "Detalle Biodiesel Argentina.hyper"], dry)
    escribir("dashboard.json", generar_dashboard_legacy(registros),
             ["Detalle Biodiesel Argentina.hyper"], dry)
    go_sectores["sectores_sin_corte"] = list(SECTORES_SIN_CORTE)
    escribir("go_sectores.json", go_sectores,
             ["Mercado Argentino Derivados Petroleo Table.hyper"], dry)
    escribir("gestion.json", dict(
        secretarios=secretarios,
        subsecretarios=subsecretarios,
        presidencias=presidencias,
        formulas_serie=form_serie, formulas_periodos=form_periodos,
    ), ["Secretaria de Energía.hyper", "Formulas de precio.hyper"], dry)

    if WARNINGS:
        print(f"\nTerminado con {len(WARNINGS)} advertencia(s) — revisar arriba.")
    else:
        print("\nTerminado sin advertencias.")


if __name__ == "__main__":
    main()
